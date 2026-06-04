"""Evaluation runner for the end-to-end RAG pipeline."""

from __future__ import annotations

import argparse
import json
import logging
from collections.abc import Mapping
from typing import Any

from agentic_rag.evaluation.metrics import mrr_at_k, recall_at_k

logger = logging.getLogger(__name__)

BASE_REQUIRED_COLUMNS = {
    "question",
    "expected_answer",
    "ground_truth_chunk_ids",
    "is_out_of_scope",
    "rag_input",
    "rag_context",
    "bot_response",
    "bot_citations",
    "trace_url",
    "retrieved_top5_ids",
    "ground_truth_rank",
    "recall_at_5",
    "mrr_at_5",
    "citation_chunk_match",
    "guardrail_pass",
}
RAGAS_REQUIRED_COLUMNS = {
    "ragas_faithfulness",
    "ragas_answer_relevancy",
    "ragas_context_precision",
    "ragas_context_recall",
}


class EvaluationRunner:
    """Runs the RAG pipeline over a test dataset and calculates metrics."""

    def __init__(self, input_file: str, output_file: str, run_ragas: bool = False) -> None:
        self.input_file = input_file
        self.output_file = output_file
        self.run_ragas = run_ragas

    def run(self) -> None:
        """Execute the evaluation pipeline."""
        try:
            import openpyxl  # type: ignore[import-untyped]
        except ImportError as exc:
            raise RuntimeError(
                "Evaluation runner requires openpyxl. "
                "Install the evaluation extra with `uv sync --extra evaluation`."
            ) from exc

        from agentic_rag.runtime_env import load_local_env
        load_local_env()

        import os
        os.environ["AGENT_RETRIEVE_WORKERS"] = "1"

        from agentic_rag.agent.graph import run_agent
        from agentic_rag.integrations.local_pdf.providers import LocalPdfEvidenceProvider

        provider = LocalPdfEvidenceProvider.from_env()

        logger.info("Loading dataset from %s", self.input_file)
        wb = openpyxl.load_workbook(self.input_file)
        if "Evaluation" not in wb.sheetnames:
            raise ValueError("Input file must contain an 'Evaluation' sheet.")

        ws = wb["Evaluation"]

        # Determine column indices from header (row 2)
        header = {cell.value: idx for idx, cell in enumerate(ws[2], start=1) if cell.value}
        _validate_required_columns(
            header,
            required_columns=BASE_REQUIRED_COLUMNS
            | (RAGAS_REQUIRED_COLUMNS if self.run_ragas else set()),
        )

        for row_idx in range(3, ws.max_row + 1):
            question_cell = ws.cell(row=row_idx, column=header["question"])
            if not question_cell.value:
                continue

            question = str(question_cell.value).strip()
            if not question:
                continue

            # Skip if already evaluated
            bot_response_cell = ws.cell(row=row_idx, column=header["bot_response"])
            if bot_response_cell.value:
                logger.info("Skipping row %s (already evaluated)", row_idx)
                continue

            logger.info("Evaluating row %s: %s", row_idx, question)

            is_out_of_scope = _is_truthy_cell(
                ws.cell(row=row_idx, column=header["is_out_of_scope"]).value
            )

            # 1. Run Agent Pipeline
            try:
                result = run_agent(provider=provider, question=question)
                evidence_chunks = result.evidence_chunks
                generation_answer = result.answer
            except Exception:
                logger.exception("Error generating answer for row %s", row_idx)
                continue

            # Format chunks to JSON
            rag_context_json = json.dumps(
                [
                    {
                        "id": chunk.chunk.chunk_id,
                        "text": chunk.chunk.text,
                        "score": chunk.score,
                        "retriever": chunk.retriever,
                    }
                    for chunk in evidence_chunks
                ],
                ensure_ascii=False,
            )

            bot_citations_json = json.dumps(
                [c.model_dump() for c in generation_answer.citations], ensure_ascii=False
            )

            # Write pipeline output
            ws.cell(row=row_idx, column=header["rag_input"]).value = question
            ws.cell(row=row_idx, column=header["rag_context"]).value = rag_context_json
            ws.cell(row=row_idx, column=header["bot_response"]).value = generation_answer.answer
            ws.cell(row=row_idx, column=header["bot_citations"]).value = bot_citations_json
            ws.cell(row=row_idx, column=header["trace_url"]).value = "N/A (check local traces)"

            # 2. Calculate Metrics — prefer storage_chunk_id (stable, user-friendly)
            top5_ids = [_resolve_chunk_id(chunk.chunk) for chunk in evidence_chunks[:5]]
            ws.cell(row=row_idx, column=header["retrieved_top5_ids"]).value = ", ".join(top5_ids)

            gt_chunks_raw = ws.cell(row=row_idx, column=header["ground_truth_chunk_ids"]).value

            if is_out_of_scope:
                is_not_found = generation_answer.status == "not_found"
                ws.cell(row=row_idx, column=header["guardrail_pass"]).value = is_not_found
            else:
                ws.cell(row=row_idx, column=header["guardrail_pass"]).value = "N/A"

                relevant_ids = _parse_relevant_ids(gt_chunks_raw)
                if relevant_ids:
                    gt_rank = -1
                    for i, cid in enumerate(top5_ids, start=1):
                        if cid in relevant_ids:
                            gt_rank = i
                            break
                    ws.cell(row=row_idx, column=header["ground_truth_rank"]).value = gt_rank

                    recall = recall_at_k(top5_ids, relevant_ids, k=5)
                    mrr = mrr_at_k(top5_ids, relevant_ids, k=5)
                    ws.cell(row=row_idx, column=header["recall_at_5"]).value = recall
                    ws.cell(row=row_idx, column=header["mrr_at_5"]).value = mrr

                    bot_citation_ids = {
                        _resolve_chunk_id(chunk.chunk)
                        for chunk in evidence_chunks
                        if any(
                        c.chunk_id == chunk.chunk.chunk_id
                        for c in generation_answer.citations
                    )
                    }
                    has_match = bool(relevant_ids.intersection(bot_citation_ids))
                    ws.cell(row=row_idx, column=header["citation_chunk_match"]).value = has_match

            # Save after each row to avoid losing progress on crash
            wb.save(self.output_file)

        # 3. RAGAS Evaluation
        if self.run_ragas:
            from agentic_rag.evaluation.ragas_eval import run_ragas_evaluation

            logger.info("Preparing data for RAGAS evaluation...")
            eval_data = []
            row_mapping = []

            for row_idx in range(3, ws.max_row + 1):
                is_out_of_scope = _is_truthy_cell(
                    ws.cell(row=row_idx, column=header["is_out_of_scope"]).value
                )
                if is_out_of_scope:
                    continue

                question = str(ws.cell(row=row_idx, column=header["question"]).value or "")
                answer = str(ws.cell(row=row_idx, column=header["bot_response"]).value or "")
                expected_answer = str(
                    ws.cell(row=row_idx, column=header["expected_answer"]).value or ""
                )
                rag_context_raw = ws.cell(row=row_idx, column=header["rag_context"]).value

                if not question or not answer:
                    continue

                contexts: list[str] = []
                if rag_context_raw:
                    try:
                        ctx_list = json.loads(rag_context_raw)
                        contexts = [
                            str(item.get("text", "")) for item in ctx_list if isinstance(item, dict)
                        ]
                    except json.JSONDecodeError:
                        logger.warning("Skipping invalid rag_context JSON in row %s", row_idx)

                eval_data.append(
                    {
                        "question": question,
                        "answer": answer,
                        "contexts": contexts,
                        "ground_truth": expected_answer,
                    }
                )
                row_mapping.append(row_idx)

            if eval_data:
                ragas_results = run_ragas_evaluation(eval_data)
                for row_idx, scores in zip(row_mapping, ragas_results, strict=False):
                    ws.cell(row=row_idx, column=header["ragas_faithfulness"]).value = scores.get(
                        "ragas_faithfulness", 0.0
                    )
                    ws.cell(
                        row=row_idx, column=header["ragas_answer_relevancy"]
                    ).value = scores.get("ragas_answer_relevancy", 0.0)
                    ws.cell(
                        row=row_idx, column=header["ragas_context_precision"]
                    ).value = scores.get("ragas_context_precision", 0.0)
                    ws.cell(row=row_idx, column=header["ragas_context_recall"]).value = scores.get(
                        "ragas_context_recall", 0.0
                    )

        # 4. Summary row
        _write_summary(ws, header, self.run_ragas)

        logger.info("Saving evaluation results to %s", self.output_file)
        wb.save(self.output_file)


def _validate_required_columns(
    header: Mapping[Any, int],
    *,
    required_columns: set[str],
) -> None:
    missing_columns = sorted(column for column in required_columns if column not in header)
    if missing_columns:
        missing = ", ".join(missing_columns)
        raise ValueError(f"Evaluation sheet is missing required columns: {missing}")


def _is_truthy_cell(value: object) -> bool:
    if isinstance(value, bool):
        return value
    if isinstance(value, int | float):
        return value != 0
    if isinstance(value, str):
        return value.strip().lower() in {"true", "1", "yes", "y"}
    return False


def _parse_relevant_ids(value: object) -> set[str]:
    if value is None:
        return set()
    raw_value = str(value).strip()
    if not raw_value or raw_value.upper() == "NONE":
        return set()
    return {chunk_id.strip() for chunk_id in raw_value.split(",") if chunk_id.strip()}


def _write_summary(ws: Any, header: dict[str, int], include_ragas: bool) -> None:
    """Append an AVERAGE summary row below all data rows."""
    last_data_row = ws.max_row
    summary_row = last_data_row + 2

    ws.cell(row=summary_row, column=1).value = "AVERAGE"

    numeric_cols = ["recall_at_5", "mrr_at_5"]
    if include_ragas:
        numeric_cols += [
            "ragas_faithfulness",
            "ragas_answer_relevancy",
            "ragas_context_precision",
            "ragas_context_recall",
        ]

    for col_name in numeric_cols:
        if col_name not in header:
            continue
        col_idx = header[col_name]
        values = [
            ws.cell(row=r, column=col_idx).value
            for r in range(3, last_data_row + 1)
            if isinstance(ws.cell(row=r, column=col_idx).value, (int, float))
        ]
        if values:
            ws.cell(row=summary_row, column=col_idx).value = round(sum(values) / len(values), 4)


def _resolve_chunk_id(chunk: Any) -> str:
    """Prefer storage_chunk_id from metadata — stable, human-readable, easy to fill in Excel."""
    storage_id = chunk.metadata.get("storage_chunk_id")
    if isinstance(storage_id, str) and storage_id:
        return storage_id
    return str(chunk.chunk_id)


def main() -> None:
    logging.basicConfig(level=logging.INFO, format="%(asctime)s %(levelname)s %(message)s")
    parser = argparse.ArgumentParser(description="Run RAG evaluation over an Excel dataset.")
    parser.add_argument("--input", required=True, help="Path to input Excel file.")
    parser.add_argument("--output", required=True, help="Path to output Excel file.")
    parser.add_argument("--ragas", action="store_true", help="Also run RAGAS LLM-as-judge metrics.")
    args = parser.parse_args()

    runner = EvaluationRunner(
        input_file=args.input,
        output_file=args.output,
        run_ragas=args.ragas,
    )
    runner.run()


if __name__ == "__main__":
    main()
