"""Eval Review router — mounted into the main FastAPI app under /eval-review."""

from __future__ import annotations

import contextlib
import os
import threading
import time
from pathlib import Path
from typing import Any, Literal
from zipfile import BadZipFile

import openpyxl
from fastapi import APIRouter, HTTPException
from pydantic import BaseModel

# ── Config ────────────────────────────────────────────────────────────────────

_PROJECT_ROOT = Path(__file__).resolve().parent.parent.parent

EXCEL_PATH = Path(
    os.environ.get(
        "EVAL_EXCEL_PATH",
        str(_PROJECT_ROOT / "guide" / "reports" / "result.xlsx"),
    )
)
REJECT_PATH = EXCEL_PATH.parent / "reject.xlsx"
BACKUP_PATH = EXCEL_PATH.parent / "result.backup.xlsx"
SHEET_NAME = "Evaluation"
HEADER_ROW = 2
DATA_START_ROW = 3
_WORKBOOK_LOAD_RETRIES = 6
_WORKBOOK_LOAD_RETRY_DELAY_SECONDS = 0.2
_TRANSIENT_WORKBOOK_ERRORS = (EOFError, BadZipFile, PermissionError)

# ── Router ────────────────────────────────────────────────────────────────────

router = APIRouter(tags=["eval-review"])

_file_lock = threading.Lock()

# Rows cache keyed on file mtime — auto-invalidates whenever the xlsx is
# written (by any endpoint OR the background eval worker), so it can never
# return stale data. Repeat reads skip the ~1.5s openpyxl parse.
_rows_cache: dict[str, Any] = {"mtime": None, "rows": None, "last_good": None}

# Columns the frontend list view actually consumes. The xlsx has 33 columns
# but 25 are empty pre-eval; sending them as "col": null per row bloats the
# payload to 2.2MB. slim=true returns only these → ~0.55MB.
_SLIM_FIELDS = (
    "id",
    "question",
    "expected_answer",
    "ground_truth_chunk_ids",
    "is_out_of_scope",
    "review_status",
    "display_status",
)


def _invalidate_rows_cache() -> None:
    _rows_cache["mtime"] = None
    _rows_cache["rows"] = None


# ── Per-row eval queue (Excel-backed) ────────────────────────────────────────
# eval_queue.xlsx is the single source of truth. Worker polls it every 3s.
# Columns: excel_row | force | status | enqueued_at
QUEUE_XLSX_PATH = EXCEL_PATH.parent / "eval_queue.xlsx"
_QUEUE_SHEET = "Queue"
_QUEUE_COLS = ["excel_row", "force", "status", "enqueued_at"]

_row_status: dict[int, dict[str, Any]] = {}  # fast in-memory cache for status API
_queue_lock = threading.Lock()  # protects eval_queue.xlsx
_worker_lock = threading.Lock()  # protects _worker_thread
_worker_thread: threading.Thread | None = None


def _queue_load_or_create() -> openpyxl.Workbook:
    if not QUEUE_XLSX_PATH.exists():
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = _QUEUE_SHEET
        ws.append(_QUEUE_COLS)
        wb.save(QUEUE_XLSX_PATH)
        return wb
    return openpyxl.load_workbook(QUEUE_XLSX_PATH)


def _queue_enqueue(excel_row: int, force: bool) -> None:
    """Append row to queue xlsx if not already pending/running."""
    with _queue_lock:
        wb = _queue_load_or_create()
        ws = wb[_QUEUE_SHEET]
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0] == excel_row and row[2] in ("pending", "running"):
                wb.close()
                return
        pos = ws.max_row  # current data rows = position in queue
        ws.append([excel_row, force, "pending", time.strftime("%Y-%m-%d %H:%M:%S")])
        wb.save(QUEUE_XLSX_PATH)
        wb.close()
    _row_status[excel_row] = {
        "status": "queued",
        "message": f"Đang chờ — vị trí #{pos} trong hàng đợi",
    }


def _queue_next_pending() -> tuple[int, bool] | None:
    """Mark the oldest pending row as running and return (excel_row, force)."""
    with _queue_lock:
        if not QUEUE_XLSX_PATH.exists():
            return None
        wb = openpyxl.load_workbook(QUEUE_XLSX_PATH)
        ws = wb[_QUEUE_SHEET]
        for row_idx in range(2, ws.max_row + 1):
            if ws.cell(row_idx, 3).value == "pending":
                excel_row = ws.cell(row_idx, 1).value
                force = bool(ws.cell(row_idx, 2).value)
                ws.cell(row_idx, 3).value = "running"
                wb.save(QUEUE_XLSX_PATH)
                wb.close()
                return int(excel_row), force
        wb.close()
        return None


def _queue_remove(excel_row: int) -> None:
    """Delete all entries for excel_row from queue xlsx."""
    with _queue_lock:
        if not QUEUE_XLSX_PATH.exists():
            return
        wb = openpyxl.load_workbook(QUEUE_XLSX_PATH)
        ws = wb[_QUEUE_SHEET]
        for r in range(ws.max_row, 1, -1):
            if ws.cell(r, 1).value == excel_row:
                ws.delete_rows(r)
        wb.save(QUEUE_XLSX_PATH)
        wb.close()


def _queue_reset_running_on_startup() -> None:
    """Reset rows stuck in 'running' back to 'pending' after crash/restart."""
    if not QUEUE_XLSX_PATH.exists():
        return
    try:
        wb = openpyxl.load_workbook(QUEUE_XLSX_PATH)
        ws = wb[_QUEUE_SHEET]
        pos = 1
        changed = False
        for row_idx in range(2, ws.max_row + 1):
            status = ws.cell(row_idx, 3).value
            excel_row = ws.cell(row_idx, 1).value
            if status == "running":
                ws.cell(row_idx, 3).value = "pending"
                changed = True
            if excel_row and status in ("pending", "running"):
                _row_status[int(excel_row)] = {
                    "status": "queued",
                    "message": f"Đang chờ (khôi phục sau restart) — vị trí #{pos}",
                }
                pos += 1
        if changed:
            wb.save(QUEUE_XLSX_PATH)
        wb.close()
    except Exception:
        pass


_queue_reset_running_on_startup()

_job: dict[str, Any] = {
    "status": "idle",
    "progress": 0,
    "total": 0,
    "errors": [],
    "message": "",
}

# ── Models ────────────────────────────────────────────────────────────────────


class RowUpdate(BaseModel):
    question: str | None = None
    expected_answer: str | None = None
    ground_truth_chunk_ids: str | None = None
    is_out_of_scope: bool | None = None


class RunConfig(BaseModel):
    run_ragas: bool = True


class JobStatus(BaseModel):
    status: Literal["idle", "running", "done", "error"]
    progress: int
    total: int
    errors: list[str]
    message: str


# ── Excel helpers ─────────────────────────────────────────────────────────────


def _load_wb() -> openpyxl.Workbook:
    if not EXCEL_PATH.exists():
        raise HTTPException(status_code=404, detail=f"Excel not found: {EXCEL_PATH}")

    last_error: BaseException | None = None
    for attempt in range(_WORKBOOK_LOAD_RETRIES):
        try:
            return openpyxl.load_workbook(EXCEL_PATH)
        except _TRANSIENT_WORKBOOK_ERRORS as exc:
            last_error = exc
            if attempt < _WORKBOOK_LOAD_RETRIES - 1:
                time.sleep(_WORKBOOK_LOAD_RETRY_DELAY_SECONDS * (attempt + 1))

    raise HTTPException(
        status_code=503,
        detail="Excel is being written. Please retry shortly.",
    ) from last_error


def _save_wb(wb: openpyxl.Workbook) -> None:
    """Save main workbook then overwrite the rolling backup copy."""
    import shutil

    wb.save(EXCEL_PATH)
    with contextlib.suppress(Exception):
        shutil.copy2(EXCEL_PATH, BACKUP_PATH)


def _read_header(ws: Any) -> dict[str, int]:
    return {cell.value: idx for idx, cell in enumerate(ws[HEADER_ROW], start=1) if cell.value}


def _ensure_review_status_col(ws: Any, header: dict[str, int]) -> bool:
    """Return True if column was added (caller should save)."""
    if "review_status" in header:
        return False
    next_col = max(header.values()) + 1
    ws.cell(row=HEADER_ROW, column=next_col).value = "review_status"
    header["review_status"] = next_col
    q_col = header.get("question", 1)
    for row_idx in range(DATA_START_ROW, ws.max_row + 1):
        if ws.cell(row=row_idx, column=q_col).value:
            ws.cell(row=row_idx, column=next_col).value = "pending"
    return True


def _row_to_dict(ws: Any, excel_row: int, header: dict[str, int]) -> dict[str, Any]:
    row: dict[str, Any] = {"excel_row": excel_row}
    for col_name, col_idx in header.items():
        row[col_name] = ws.cell(row=excel_row, column=col_idx).value

    review_status = row.get("review_status") or "pending"
    bot_response = row.get("bot_response")
    if bot_response:
        row["display_status"] = "evaluated"
    else:
        row["display_status"] = review_status

    return row


def _question_id_number(value: Any) -> int | None:
    text = str(value or "").strip()
    digits = "".join(ch for ch in text if ch.isdigit())
    return int(digits) if digits else None


def _read_all_rows() -> list[dict[str, Any]]:
    with _file_lock:
        mtime = EXCEL_PATH.stat().st_mtime if EXCEL_PATH.exists() else None
        if mtime is not None and _rows_cache["mtime"] == mtime and _rows_cache["rows"] is not None:
            return _rows_cache["rows"]  # type: ignore[no-any-return]

        try:
            wb = _load_wb()
        except HTTPException as exc:
            if exc.status_code == 503:
                fallback = _rows_cache["rows"] or _rows_cache["last_good"]
                if fallback is not None:
                    return fallback  # type: ignore[no-any-return]
            raise
        ws = wb[SHEET_NAME] if SHEET_NAME in wb.sheetnames else wb.active
        header = _read_header(ws)

        col_added = _ensure_review_status_col(ws, header)
        if col_added:
            _save_wb(wb)

        q_col = header.get("question", 1)
        rows = []
        for excel_row in range(DATA_START_ROW, ws.max_row + 1):
            if ws.cell(row=excel_row, column=q_col).value:
                rows.append(_row_to_dict(ws, excel_row, header))

        wb.close()

        # Re-stat: _ensure_review_status_col may have re-saved the file above.
        _rows_cache["mtime"] = EXCEL_PATH.stat().st_mtime if EXCEL_PATH.exists() else None
        _rows_cache["rows"] = rows
        _rows_cache["last_good"] = rows  # never cleared — fallback on 503
        return rows


# ── Routes ────────────────────────────────────────────────────────────────────


@router.get("/api/rows")
def get_rows(slim: bool = False) -> list[dict[str, Any]]:
    rows = _read_all_rows()
    if slim:
        return [{"excel_row": r["excel_row"], **{k: r.get(k) for k in _SLIM_FIELDS}} for r in rows]
    return rows


@router.patch("/api/rows/{excel_row}")
def update_row(excel_row: int, update: RowUpdate) -> dict[str, Any]:
    with _file_lock:
        wb = _load_wb()
        ws = wb[SHEET_NAME] if SHEET_NAME in wb.sheetnames else wb.active
        header = _read_header(ws)

        q_col = header.get("question", 1)
        if not ws.cell(row=excel_row, column=q_col).value:
            raise HTTPException(status_code=404, detail="Row not found")

        for field, value in update.model_dump(exclude_none=True).items():
            if field in header:
                ws.cell(row=excel_row, column=header[field]).value = value

        _save_wb(wb)
        _invalidate_rows_cache()
        result = _row_to_dict(ws, excel_row, header)
        wb.close()
        return result


@router.post("/api/rows/{excel_row}/approve")
def approve_row(excel_row: int) -> dict[str, Any]:
    with _file_lock:
        wb = _load_wb()
        ws = wb[SHEET_NAME] if SHEET_NAME in wb.sheetnames else wb.active
        header = _read_header(ws)
        _ensure_review_status_col(ws, header)

        ws.cell(row=excel_row, column=header["review_status"]).value = "approved"
        _save_wb(wb)
        _invalidate_rows_cache()
        result = _row_to_dict(ws, excel_row, header)
        wb.close()
        return result


@router.post("/api/rows/{excel_row}/approve-and-eval")
def approve_and_eval_row(excel_row: int) -> dict[str, Any]:
    """Approve a row and enqueue it for pipeline evaluation. Returns immediately."""
    with _file_lock:
        wb = _load_wb()
        ws = wb[SHEET_NAME] if SHEET_NAME in wb.sheetnames else wb.active
        header = _read_header(ws)
        _ensure_review_status_col(ws, header)

        q_col = header.get("question", 1)
        if not ws.cell(row=excel_row, column=q_col).value:
            raise HTTPException(status_code=404, detail="Row not found")

        ws.cell(row=excel_row, column=header["review_status"]).value = "approved"
        _save_wb(wb)
        _invalidate_rows_cache()
        row_dict = _row_to_dict(ws, excel_row, header)
        wb.close()

    _queue_enqueue(excel_row, False)
    _ensure_eval_worker()
    return {**row_dict, "_eval_status": _row_status[excel_row]}


@router.post("/api/rows/{excel_row}/re-eval")
def re_eval_row(excel_row: int) -> dict[str, Any]:
    """Re-evaluate an already-evaluated row (force=True bypasses the skip check)."""
    with _file_lock:
        wb = _load_wb()
        ws = wb[SHEET_NAME] if SHEET_NAME in wb.sheetnames else wb.active
        header = _read_header(ws)
        q_col = header.get("question", 1)
        if not ws.cell(row=excel_row, column=q_col).value:
            raise HTTPException(status_code=404, detail="Row not found")
        row_dict = _row_to_dict(ws, excel_row, header)
        wb.close()

    _queue_enqueue(excel_row, True)
    _ensure_eval_worker()
    return {**row_dict, "_eval_status": _row_status[excel_row]}


@router.get("/api/rows/{excel_row}/eval-status")
def get_row_eval_status(excel_row: int) -> dict[str, Any]:
    return _row_status.get(excel_row, {"status": "idle", "message": ""})


@router.get("/api/eval/active-status")
def get_active_eval_status() -> dict[str, Any]:
    """Return status of all rows currently queued/running — one call replaces N polls."""
    return {
        str(row): status
        for row, status in _row_status.items()
        if status.get("status") in {"queued", "running", "done", "error"}
    }


# ── Queue worker ──────────────────────────────────────────────────────────────

_WORKER_POLL_INTERVAL = 3  # seconds between queue polls


def _ensure_eval_worker() -> None:
    global _worker_thread
    with _worker_lock:
        if _worker_thread is None or not _worker_thread.is_alive():
            _worker_thread = threading.Thread(target=_eval_worker_loop, daemon=True)
            _worker_thread.start()


def _eval_worker_loop() -> None:
    """Poll eval_queue.xlsx every 3s and process rows one at a time."""
    while True:
        item = _queue_next_pending()
        if item is None:
            time.sleep(_WORKER_POLL_INTERVAL)
            continue
        excel_row, force = item
        try:
            _row_status[excel_row] = {"status": "running", "message": "Đang chạy pipeline..."}
            _run_single_row_eval(excel_row, force=force)
            _invalidate_rows_cache()
            _row_status[excel_row] = {"status": "done", "message": "Hoàn thành"}
        except Exception as exc:
            _row_status[excel_row] = {"status": "error", "message": str(exc)}
        finally:
            _queue_remove(excel_row)


# Start worker after all functions are defined
threading.Thread(target=_eval_worker_loop, daemon=True).start()


_EVAL_OUTPUT_COLS = {
    "rag_input",
    "rag_context",
    "bot_response",
    "bot_citations",
    "trace_url",
    "retrieved_top5_ids",
    "ground_truth_rank",
    "recall_at_5",
    "mrr_at_5",
    "citation_chunk_match",
    "guardrail_pass",
    "ragas_faithfulness",
    "ragas_answer_relevancy",
    "ragas_context_precision",
    "ragas_context_recall",
}


def _run_single_row_eval(excel_row: int, *, force: bool = False) -> None:
    """Run EvaluationRunner on a temp copy, then merge only eval columns back under lock."""
    import shutil

    from agentic_rag.runtime_env import load_local_env

    load_local_env()

    from agentic_rag.evaluation.runner import EvaluationRunner

    tmp_path = EXCEL_PATH.with_suffix(".eval_tmp.xlsx")

    # 1. Snapshot current file (fast, under lock)
    with _file_lock:
        shutil.copy2(EXCEL_PATH, tmp_path)

    # 2. Run eval on temp copy — never touches result.xlsx
    try:
        EvaluationRunner(
            input_file=str(tmp_path),
            output_file=str(tmp_path),
            run_ragas=True,
            target_row=excel_row,
            force=force,
            on_row_done=lambda: _row_status[excel_row].update({"message": "Đang ghi kết quả..."}),
        ).run()
    except Exception:
        tmp_path.unlink(missing_ok=True)
        raise

    # 3. Merge only eval output columns back into result.xlsx (fast, under lock)
    with _file_lock:
        tmp_wb = openpyxl.load_workbook(tmp_path)
        tmp_ws = tmp_wb[SHEET_NAME] if SHEET_NAME in tmp_wb.sheetnames else tmp_wb.active
        tmp_header = _read_header(tmp_ws)

        main_wb = _load_wb()
        main_ws = main_wb[SHEET_NAME] if SHEET_NAME in main_wb.sheetnames else main_wb.active
        main_header = _read_header(main_ws)

        for col_name in _EVAL_OUTPUT_COLS:
            if col_name in tmp_header and col_name in main_header:
                val = tmp_ws.cell(row=excel_row, column=tmp_header[col_name]).value
                main_ws.cell(row=excel_row, column=main_header[col_name]).value = val

        _save_wb(main_wb)
        main_wb.close()
        tmp_wb.close()

    tmp_path.unlink(missing_ok=True)
    _invalidate_rows_cache()


@router.post("/api/rows/{excel_row}/reject")
def reject_row(excel_row: int) -> dict[str, Any]:
    with _file_lock:
        wb = _load_wb()
        ws = wb[SHEET_NAME] if SHEET_NAME in wb.sheetnames else wb.active
        header = _read_header(ws)

        q_col = header.get("question", 1)
        question = ws.cell(row=excel_row, column=q_col).value
        if not question:
            raise HTTPException(status_code=404, detail="Row not found")

        row_data = {
            col_name: ws.cell(row=excel_row, column=col_idx).value
            for col_name, col_idx in header.items()
        }
        _append_to_reject(row_data, header)

        ws.delete_rows(excel_row)
        _save_wb(wb)
        _invalidate_rows_cache()
        wb.close()

        return {"ok": True, "deleted_question": str(question)}


@router.get("/api/rejected")
def get_rejected_rows() -> list[dict[str, Any]]:
    if not REJECT_PATH.exists():
        return []

    with _file_lock:
        rb = openpyxl.load_workbook(REJECT_PATH, read_only=True, data_only=True)
        rws = rb["Rejected"] if "Rejected" in rb.sheetnames else rb.active
        header = {cell.value: idx for idx, cell in enumerate(rws[1], start=1) if cell.value}
        q_col = header.get("question", 1)
        rows = []
        for reject_row in range(2, rws.max_row + 1):
            if not rws.cell(row=reject_row, column=q_col).value:
                continue
            row: dict[str, Any] = {"reject_row": reject_row, "excel_row": reject_row}
            for col_name, col_idx in header.items():
                row[col_name] = rws.cell(row=reject_row, column=col_idx).value
            row["display_status"] = "deleted"
            rows.append(row)
        rb.close()
        return rows


@router.post("/api/rejected/{reject_row}/restore")
def restore_rejected_row(reject_row: int) -> dict[str, Any]:
    if reject_row < 2:
        raise HTTPException(status_code=404, detail="Rejected row not found")
    if not REJECT_PATH.exists():
        raise HTTPException(status_code=404, detail="Rejected workbook not found")

    with _file_lock:
        rb = openpyxl.load_workbook(REJECT_PATH)
        rws = rb["Rejected"] if "Rejected" in rb.sheetnames else rb.active
        if reject_row > rws.max_row:
            rb.close()
            raise HTTPException(status_code=404, detail="Rejected row not found")

        reject_header = {cell.value: idx for idx, cell in enumerate(rws[1], start=1) if cell.value}
        q_col = reject_header.get("question", 1)
        if not rws.cell(row=reject_row, column=q_col).value:
            rb.close()
            raise HTTPException(status_code=404, detail="Rejected row not found")

        row_data = {
            col_name: rws.cell(row=reject_row, column=col_idx).value
            for col_name, col_idx in reject_header.items()
        }
        restore_id = row_data.get("id")
        restore_id_number = _question_id_number(restore_id)

        wb = _load_wb()
        ws = wb[SHEET_NAME] if SHEET_NAME in wb.sheetnames else wb.active
        header = _read_header(ws)

        if "id" in header and restore_id:
            for row_idx in range(DATA_START_ROW, ws.max_row + 1):
                if ws.cell(row=row_idx, column=header["id"]).value == restore_id:
                    # ID already in main sheet — just remove from rejected and return existing row
                    existing = _row_to_dict(ws, row_idx, header)
                    rws.delete_rows(reject_row, 1)
                    rb.save(REJECT_PATH)
                    wb.close()
                    rb.close()
                    _invalidate_rows_cache()
                    return {"ok": True, "restored_row": existing}

        insert_at = ws.max_row + 1
        if "id" in header and restore_id_number is not None:
            for row_idx in range(DATA_START_ROW, ws.max_row + 1):
                current_id_number = _question_id_number(
                    ws.cell(row=row_idx, column=header["id"]).value
                )
                if current_id_number is not None and current_id_number > restore_id_number:
                    insert_at = row_idx
                    break

        ws.insert_rows(insert_at, 1)
        for col_name, col_idx in header.items():
            ws.cell(row=insert_at, column=col_idx).value = row_data.get(col_name)

        rws.delete_rows(reject_row, 1)
        _save_wb(wb)
        rb.save(REJECT_PATH)
        _invalidate_rows_cache()

        result = _row_to_dict(ws, insert_at, header)
        wb.close()
        rb.close()
        return {"ok": True, "restored_row": result}


def _append_to_reject(row_data: dict[str, Any], header: dict[str, int]) -> None:
    col_names = [k for k, _ in sorted(header.items(), key=lambda x: x[1])]

    if REJECT_PATH.exists():
        rb = openpyxl.load_workbook(REJECT_PATH)
        rws = rb["Rejected"] if "Rejected" in rb.sheetnames else rb.active
    else:
        rb = openpyxl.Workbook()
        rws = rb.active
        rws.title = "Rejected"
        for col_idx, col_name in enumerate(col_names, start=1):
            rws.cell(row=1, column=col_idx).value = col_name

    next_row = rws.max_row + 1
    for col_idx, col_name in enumerate(col_names, start=1):
        rws.cell(row=next_row, column=col_idx).value = row_data.get(col_name)

    rb.save(REJECT_PATH)
    rb.close()


@router.post("/api/eval/run")
def run_eval(config: RunConfig) -> dict[str, Any]:
    global _job
    if _job["status"] == "running":
        raise HTTPException(status_code=409, detail="Evaluation already running")

    _job = {
        "status": "running",
        "progress": 0,
        "total": 0,
        "errors": [],
        "message": "Starting...",
    }

    thread = threading.Thread(target=_run_eval_job, args=(config.run_ragas,), daemon=True)
    thread.start()
    return {"message": "Evaluation started"}


def _run_eval_job(run_ragas: bool) -> None:
    global _job
    try:
        from agentic_rag.evaluation.runner import EvaluationRunner

        _job["message"] = "Counting approved rows..."
        rows = _read_all_rows()
        approved = [
            r for r in rows if r.get("review_status") == "approved" and not r.get("bot_response")
        ]
        _job["total"] = len(approved)

        if not approved:
            _job["status"] = "done"
            _job["message"] = "No approved rows to evaluate"
            return

        _job["message"] = f"Running on {len(approved)} approved rows..."

        runner = EvaluationRunner(
            input_file=str(EXCEL_PATH),
            output_file=str(EXCEL_PATH),
            run_ragas=run_ragas,
            approved_only=True,
            on_row_done=_increment_progress,
        )
        runner.run()
        _invalidate_rows_cache()

        _job["status"] = "done"
        _job["message"] = f"Done! {_job['progress']}/{_job['total']} rows evaluated."

    except Exception as exc:
        _job["status"] = "error"
        _job["errors"].append(str(exc))
        _job["message"] = f"Error: {exc}"


def _increment_progress() -> None:
    _job["progress"] += 1


@router.get("/api/eval/status")
def get_eval_status() -> JobStatus:
    return JobStatus(**_job)


_qdrant_cache: dict[str, Any] = {"at": 0.0, "payloads": []}
_qdrant_cache_lock = threading.Lock()
_QDRANT_CACHE_TTL = 60  # seconds


def _all_qdrant_payloads(client: Any, collection_name: str) -> list[dict[str, Any]]:
    """Return every chunk payload in the collection, cached for _QDRANT_CACHE_TTL.

    document_id in Qdrant is a random local_url_<uuid>; the deterministic
    url_<hash> prefix lives only inside storage_chunk_id. There is no payload
    index that lets us prefix-filter server-side, so we scroll once and reuse.
    """
    now = time.monotonic()
    with _qdrant_cache_lock:
        if _qdrant_cache["payloads"] and (now - _qdrant_cache["at"]) < _QDRANT_CACHE_TTL:
            return _qdrant_cache["payloads"]  # type: ignore[no-any-return]

    payloads: list[dict[str, Any]] = []
    offset = None
    while True:
        batch, offset = client.scroll(
            collection_name=collection_name,
            limit=250,
            offset=offset,
            with_payload=True,
            with_vectors=False,
        )
        payloads.extend(p.payload for p in batch if isinstance(p.payload, dict))
        if offset is None:
            break

    with _qdrant_cache_lock:
        _qdrant_cache["payloads"] = payloads
        _qdrant_cache["at"] = time.monotonic()
    return payloads


@router.get("/api/doc-chunks")
def get_doc_chunks(chunk_id: str) -> dict[str, Any]:
    """Query vector store metadata for all chunks from the same document.

    Supports both pgvector (langchain_pg_embedding) and Qdrant Cloud backends,
    selected automatically from DENSE_VECTOR_STORE env var.
    """
    import os
    import re

    from agentic_rag.runtime_env import load_local_env

    load_local_env()

    # chunk_id = "url_9dd8e94fee25_section_c001" → doc_prefix = "url_9dd8e94fee25"
    parts = chunk_id.split("_")
    if len(parts) < 2:
        raise HTTPException(status_code=400, detail="Invalid chunk_id format")
    doc_prefix = f"{parts[0]}_{parts[1]}"

    def _chunk_index(cid: str) -> int:
        m = re.search(r"_c(\d+)$", cid)
        return int(m.group(1)) if m else 0

    vector_store = os.environ.get("DENSE_VECTOR_STORE", "pgvector").lower()

    # ── Qdrant path ───────────────────────────────────────────────────────────
    if vector_store == "qdrant":
        qdrant_url = os.environ.get("QDRANT_URL", "")
        qdrant_api_key = os.environ.get("QDRANT_API_KEY")
        collection_name = os.environ.get("QDRANT_COLLECTION", "agentic_rag_chunks")

        if not qdrant_url:
            raise HTTPException(status_code=503, detail="QDRANT_URL not set")

        try:
            from qdrant_client import QdrantClient
            from qdrant_client import models as qmodels
        except ImportError as exc:
            raise HTTPException(status_code=503, detail="qdrant-client not installed") from exc

        try:
            client = QdrantClient(url=qdrant_url, api_key=qdrant_api_key)
            # Fast path: with the deterministic-id backend, document_id == doc_prefix
            # (url_<hash>), so we can filter server-side by the indexed document_id.
            points, _ = client.scroll(
                collection_name=collection_name,
                scroll_filter=qmodels.Filter(
                    must=[
                        qmodels.FieldCondition(
                            key="document_id",
                            match=qmodels.MatchValue(value=doc_prefix),
                        )
                    ]
                ),
                limit=500,
                with_payload=True,
                with_vectors=False,
            )
            matched = [p.payload for p in points if isinstance(p.payload, dict)]
            if not matched:
                # Fallback for legacy docs whose document_id is a random local_url_<uuid>:
                # scan the cached payloads and prefix-match storage_chunk_id.
                prefix = f"{doc_prefix}_"
                matched = [
                    p
                    for p in _all_qdrant_payloads(client, collection_name)
                    if str(p.get("storage_chunk_id") or p.get("chunk_id") or "").startswith(prefix)
                ]
        except Exception as exc:
            raise HTTPException(status_code=503, detail=f"Qdrant query failed: {exc}") from exc

        if not matched:
            return {"document_id": doc_prefix, "found": False, "chunks": []}

        chunks = sorted(
            [
                {
                    "chunk_id": str(p.get("storage_chunk_id") or p.get("chunk_id") or ""),
                    "text": str(p.get("text") or ""),
                    "score": 1.0,
                    "retriever": "",
                    "section": str(p.get("section") or ""),
                    "url": str(p.get("url") or ""),
                }
                for p in matched
            ],
            key=lambda c: _chunk_index(str(c["chunk_id"])),
        )
        return {"document_id": doc_prefix, "found": True, "chunks": chunks}

    # ── pgvector path (default) ───────────────────────────────────────────────
    import psycopg

    conn_str = os.environ.get("DENSE_PGVECTOR_CONNECTION", "")
    collection_name = os.environ.get("DENSE_PGVECTOR_COLLECTION", "agentic_rag_chunks")

    if not conn_str:
        raise HTTPException(status_code=503, detail="DENSE_PGVECTOR_CONNECTION not set")

    db_url = re.sub(r"\+\w+://", "://", conn_str)  # postgresql+psycopg:// → postgresql://

    try:
        with psycopg.connect(db_url) as conn, conn.cursor() as cur:
            cur.execute(
                """
                SELECT DISTINCT ON (e.cmetadata->>'chunk_id') e.document, e.cmetadata
                FROM langchain_pg_embedding e
                JOIN langchain_pg_collection c ON e.collection_id = c.uuid
                WHERE c.name = %s
                  AND starts_with(e.cmetadata->>'chunk_id', %s)
                ORDER BY e.cmetadata->>'chunk_id'
                """,
                (collection_name, f"{doc_prefix}_"),
            )
            rows = cur.fetchall()
    except Exception as exc:
        raise HTTPException(status_code=503, detail=f"pgvector query failed: {exc}") from exc

    if not rows:
        return {"document_id": doc_prefix, "found": False, "chunks": []}

    def _parse_meta(raw: Any) -> dict[str, Any]:
        if isinstance(raw, str):
            import json as _json

            return _json.loads(raw)  # type: ignore[no-any-return]
        return raw or {}

    chunks = sorted(
        [
            {
                "chunk_id": _parse_meta(meta).get("chunk_id", ""),
                "text": text,
                "score": 1.0,
                "retriever": "",
                "section": _parse_meta(meta).get("section", ""),
                "url": _parse_meta(meta).get("url", ""),
            }
            for text, meta in rows
        ],
        key=lambda c: _chunk_index(str(c["chunk_id"])),
    )
    return {"document_id": doc_prefix, "found": True, "chunks": chunks}


@router.get("/api/chunks")
def get_chunks(q: str) -> list[dict[str, Any]]:
    try:
        from agentic_rag.runtime_env import load_local_env

        load_local_env()

        from agentic_rag.generation.evidence import evidence_for_question

        results, _ = evidence_for_question(question=q)
        return [
            {
                "chunk_id": r.chunk.chunk_id,
                "text": r.chunk.text[:800],
                "score": round(float(r.score), 4),
                "retriever": getattr(r, "retriever", ""),
                "section": r.chunk.metadata.get("section", ""),
                "url": r.chunk.metadata.get("url") or "",
            }
            for r in results[:10]
        ]
    except Exception as exc:
        raise HTTPException(status_code=503, detail=f"Retrieval unavailable: {exc}") from exc
