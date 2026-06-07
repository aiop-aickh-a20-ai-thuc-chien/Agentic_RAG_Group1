import openpyxl, sys
from pathlib import Path
sys.stdout.reconfigure(encoding="utf-8")

src = Path("guide/reports/result_final.xlsx")
wb = openpyxl.load_workbook(src)
ws = wb.active

headers = [c.value for c in next(ws.iter_rows(min_row=2, max_row=2))]
col = {h: i+1 for i, h in enumerate(headers) if h}

numeric_cols = ["recall_at_5","mrr_at_5","citation_chunk_match",
                "ragas_faithfulness","ragas_answer_relevancy","ragas_context_precision","ragas_context_recall"]

vals = {c: [] for c in numeric_cols}
total = 0; in_scope = 0; out_scope = 0

for row in ws.iter_rows(min_row=3, values_only=True):
    if not any(row): continue
    total += 1
    oos = row[col.get("is_out_of_scope", 1) - 1]
    if str(oos).upper() == "TRUE":
        out_scope += 1
    else:
        in_scope += 1
    for c in numeric_cols:
        v = row[col[c] - 1] if c in col else None
        if v is not None:
            try:
                vals[c].append(float(v))
            except Exception:
                pass

# AVERAGE row
avg_row = ["AVERAGE"] + [None] * (len(headers) - 1)
for c in numeric_cols:
    if c in col and vals[c]:
        avg_row[col[c] - 1] = round(sum(vals[c]) / len(vals[c]), 4)
ws.append(avg_row)

def avg(lst):
    return round(sum(lst) / len(lst), 4) if lst else None

# Summary sheet
if "Summary" in wb.sheetnames:
    del wb["Summary"]
ws_sum = wb.create_sheet("Summary")

summary_data = [
    ("Metric", "Giá trị", "Mục tiêu", "Ghi chú"),
    ("Tổng câu hỏi test", total, ">= 10", "Đếm số dòng có ID"),
    ("Câu in-scope", in_scope, "", "is_out_of_scope = FALSE"),
    ("Câu out-of-scope", out_scope, ">= 1", "is_out_of_scope = TRUE"),
    ("", "", "", ""),
    ("Recall@5 (trung bình)", avg(vals["recall_at_5"]), ">= 0.70", "Trung bình recall_at_5"),
    ("MRR@5 (trung bình)", avg(vals["mrr_at_5"]), ">= 0.50", "Trung bình mrr_at_5"),
    ("", "", "", ""),
    ("Citation Accuracy", avg(vals["citation_chunk_match"]), ">= 0.80", "Tỉ lệ citation khớp chunk"),
    ("", "", "", ""),
    ("RAGAS Faithfulness", avg(vals["ragas_faithfulness"]), ">= 0.80", "Độ trung thực answer so với context"),
    ("RAGAS Answer Relevancy", avg(vals["ragas_answer_relevancy"]), ">= 0.80", "Độ liên quan answer so với câu hỏi"),
    ("RAGAS Context Precision", avg(vals["ragas_context_precision"]), ">= 0.80", "Tỉ lệ context hữu ích"),
    ("RAGAS Context Recall", avg(vals["ragas_context_recall"]), ">= 0.80", "Khả năng retrieve đủ thông tin"),
]

ws_sum.append(("📊 Evaluation Summary (1002 rows)",))
ws_sum.append(())
for row in summary_data:
    ws_sum.append(row)

wb.save(src)
wb.close()

print("Done")
print(f"Total: {total} | In-scope: {in_scope} | Out-scope: {out_scope}")
for c in numeric_cols:
    print(f"  {c}: {avg(vals[c])}")
