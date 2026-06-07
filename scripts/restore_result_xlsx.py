"""Restore guide/reports/result.xlsx from backup_qa.tsv."""

from __future__ import annotations

import argparse
import csv
from pathlib import Path

import openpyxl  # type: ignore[import-untyped]
from openpyxl.styles import Alignment, Font, PatternFill  # type: ignore[import-untyped]
from openpyxl.utils import get_column_letter  # type: ignore[import-untyped]


COLUMNS = [
    "id",                      # 1
    "section_name",            # 2
    "question",                # 3
    "expected_answer",         # 4
    "ground_truth_chunk_ids",  # 5
    "ground_truth_doc",        # 6
    "ground_truth_page",       # 7
    "is_out_of_scope",         # 8
    "custom_preconds",         # 9
    "rag_input",               # 10
    "rag_context",             # 11
    "bot_response",            # 12
    "bot_citations",           # 13
    "trace_url",               # 14
    "retrieved_top5_ids",      # 15
    "ground_truth_rank",       # 16
    "recall_at_5",             # 17
    "mrr_at_5",                # 18
    "citation_chunk_match",    # 19
    "guardrail_pass",          # 20
    "check_answer_correct",    # 21
    "check_answer_reason",     # 22
    "check_kb_used",           # 23
    "check_kb_reason",         # 24
    "check_citation_correct",  # 25
    "check_citation_reason",   # 26
    "error_type",              # 27
    "overall_verdict",         # 28
    "ragas_faithfulness",      # 29
    "ragas_answer_relevancy",  # 30
    "ragas_context_precision", # 31
    "ragas_context_recall",    # 32
    "review_status",           # 33
]

# (col_start, col_end, label, hex_color)
SECTIONS = [
    (1,  9,  "🏷️ HUMAN: Test Dataset",   "4CAF50"),
    (10, 14, "🤖 CODE: Pipeline Output",  "1565C0"),
    (15, 20, "🤖 CODE: Auto Metrics",     "1565C0"),
    (21, 28, "👤 HUMAN: Review",          "E65100"),
    (29, 33, "🤖 LLM Judge: RAGAS",       "6A1B9A"),
]


def _hex_fill(hex_color: str) -> PatternFill:
    return PatternFill("solid", fgColor=hex_color)


def restore(tsv_path: str, out_path: str) -> None:
    print(f"Reading TSV: {tsv_path}")
    rows: list[dict[str, str]] = []
    with open(tsv_path, encoding="utf-8", newline="") as f:
        reader = csv.DictReader(f, delimiter="\t")
        for row in reader:
            rows.append(row)
    print(f"  Loaded {len(rows)} rows")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Evaluation"

    white_bold = Font(bold=True, color="FFFFFF")
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    wrap = Alignment(wrap_text=True, vertical="top")

    # Row 1 — section headers
    ws.row_dimensions[1].height = 24
    for col_start, col_end, label, color in SECTIONS:
        cell = ws.cell(row=1, column=col_start, value=label)
        cell.font = white_bold
        cell.fill = _hex_fill(color)
        cell.alignment = center
        if col_end > col_start:
            ws.merge_cells(
                f"{get_column_letter(col_start)}1:{get_column_letter(col_end)}1"
            )

    # Row 2 — column headers
    ws.row_dimensions[2].height = 30
    header_fill = _hex_fill("212121")
    for col_idx, col_name in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=2, column=col_idx, value=col_name)
        cell.font = white_bold
        cell.fill = header_fill
        cell.alignment = center

    # Row 3+ — data
    for row_num, row in enumerate(rows, start=3):
        q_idx = row_num - 2
        ws.cell(row=row_num, column=1).value = f"Q{q_idx}"
        ws.cell(row=row_num, column=3).value = row.get("question", "")
        ws.cell(row=row_num, column=4).value = row.get("expected_answer", "")
        ws.cell(row=row_num, column=5).value = row.get("chunk_id", "")
        ws.cell(row=row_num, column=6).value = row.get("url", "")
        ws.cell(row=row_num, column=8).value = False
        ws.cell(row=row_num, column=33).value = "pending"
        for col_idx in range(1, len(COLUMNS) + 1):
            ws.cell(row=row_num, column=col_idx).alignment = wrap

    # Column widths
    col_widths = {
        1: 6, 2: 15, 3: 40, 4: 40, 5: 35, 6: 40, 7: 8, 8: 12, 9: 15,
        10: 40, 11: 50, 12: 50, 13: 40, 14: 30,
        15: 40, 16: 10, 17: 10, 18: 10, 19: 15, 20: 12,
        21: 12, 22: 30, 23: 12, 24: 30, 25: 15, 26: 30, 27: 15, 28: 15,
        29: 15, 30: 15, 31: 15, 32: 15,
        33: 12,
    }
    for col_idx, width in col_widths.items():
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    ws.freeze_panes = "A3"

    Path(out_path).parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
    print(f"Saved: {out_path}  ({len(rows)} data rows)")


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--tsv", required=True)
    parser.add_argument("--out", default="guide/reports/result.xlsx")
    args = parser.parse_args()
    restore(args.tsv, args.out)


if __name__ == "__main__":
    main()
