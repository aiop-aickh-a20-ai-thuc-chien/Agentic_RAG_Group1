"""
Delete 10 rows (missing chunk IDs) from result.xlsx, renumber IDs,
and remove failed URLs from _relink.txt.
"""

import io
import sys

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")

from pathlib import Path  # noqa: E402

import openpyxl  # noqa: E402

XLSX_PATH = r"C:\Users\ACER\Downloads\Agentic_RAG_Group1\guide\reports\result.xlsx"
RELINK_PATH = r"C:\Users\ACER\Downloads\Agentic_RAG_Group1\_relink.txt"
FAILED_PATH = r"C:\Users\ACER\Downloads\Agentic_RAG_Group1\scripts\upload_failed.txt"

# IDs to delete
DELETE_IDS = {
    "Q694",
    "Q792",
    "Q1488",
    "Q1489",
    "Q1493",
    "Q1495",
    "Q1499",
    "Q1506",
    "Q1915",
    "Q2095",
}

# ── Load xlsx ─────────────────────────────────────────────────────────────────
print("Loading result.xlsx ...")
wb = openpyxl.load_workbook(XLSX_PATH)
ws = wb.active

# Find id column in row 2 (header row)
header_row_idx = 2
id_col = None
for cell in ws[header_row_idx]:
    if cell.value == "id":
        id_col = cell.column
        break

if id_col is None:
    print("ERROR: 'id' column not found in row 2")
    sys.exit(1)

print(f"  'id' column is column {id_col}")

# Collect rows to delete (row numbers, 1-indexed)
rows_to_delete = []
for row_idx in range(header_row_idx + 1, ws.max_row + 1):
    val = ws.cell(row=row_idx, column=id_col).value
    if val and str(val).strip() in DELETE_IDS:
        rows_to_delete.append(row_idx)

ids = [ws.cell(r, id_col).value for r in rows_to_delete]
print(f"  Found {len(rows_to_delete)} rows to delete: {ids}")

# Delete rows in reverse order (so indices don't shift)
for row_idx in reversed(rows_to_delete):
    ws.delete_rows(row_idx)

print(f"  Deleted {len(rows_to_delete)} rows")

# ── Renumber IDs ──────────────────────────────────────────────────────────────
print("Renumbering IDs ...")
counter = 1
for row_idx in range(header_row_idx + 1, ws.max_row + 1):
    cell = ws.cell(row=row_idx, column=id_col)
    if cell.value is not None or row_idx <= ws.max_row:
        # Only renumber data rows (skip if entire row is empty)
        row_vals = [
            ws.cell(row=row_idx, column=c).value for c in range(1, min(6, ws.max_column + 1))
        ]
        if any(v is not None for v in row_vals):
            cell.value = f"Q{counter}"
            counter += 1

total_data = counter - 1
print(f"  Renumbered Q1 to Q{total_data} ({total_data} rows)")

# ── Save xlsx ─────────────────────────────────────────────────────────────────
print("Saving result.xlsx ...")
wb.save(XLSX_PATH)
print("  Saved.")

# ── Remove failed URLs from _relink.txt ───────────────────────────────────────
print("\nUpdating _relink.txt ...")
failed_urls: set[str] = set()
fp = Path(FAILED_PATH)
if fp.exists():
    for line in fp.read_text(encoding="utf-8").splitlines():
        line = line.strip()
        if line:
            failed_urls.add(line)

print(f"  URLs to remove: {len(failed_urls)}")
for u in sorted(failed_urls):
    print(f"    {u}")

relink = Path(RELINK_PATH)
original = relink.read_text(encoding="utf-8").splitlines()
kept = [line for line in original if line.strip() not in failed_urls]
removed = len(original) - len(kept)
relink.write_text("\n".join(kept) + "\n", encoding="utf-8")
print(f"  Removed {removed} URLs, kept {len(kept)} URLs")

print("\nDone.")
