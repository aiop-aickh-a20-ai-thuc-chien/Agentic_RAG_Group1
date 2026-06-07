"""Check S3 & Qdrant coverage against result.xlsx ground_truth_chunk_ids.

Usage:
    uv run --no-sync python scripts/check_coverage.py
"""
import sys, io, os
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")

from dotenv import load_dotenv
load_dotenv()

import openpyxl
import boto3
from qdrant_client import QdrantClient
from qdrant_client.models import Filter, FieldCondition, MatchValue

XLSX_PATH = r"C:\Users\ACER\Downloads\Agentic_RAG_Group1\guide\reports\result.xlsx"

# ── 1. Load xlsx ──────────────────────────────────────────────────────────────
print("Loading result.xlsx ...")
wb = openpyxl.load_workbook(XLSX_PATH, read_only=True, data_only=True)
ws = wb.active

# Row 2 = column headers (row 1 = section headers)
header_row = [cell.value for cell in ws[2]]
col_chunk  = header_row.index("ground_truth_chunk_ids")  # col 5 (0-indexed 4)
col_doc    = header_row.index("ground_truth_doc")         # col 6 (0-indexed 5)

xlsx_chunk_ids: set[str] = set()
xlsx_urls: set[str] = set()
for row in ws.iter_rows(min_row=3, values_only=True):
    cid = row[col_chunk]
    url = row[col_doc]
    if cid and str(cid).strip():
        xlsx_chunk_ids.add(str(cid).strip())
    if url and str(url).strip().startswith("http"):
        xlsx_urls.add(str(url).strip())

print(f"  Unique chunk IDs in xlsx : {len(xlsx_chunk_ids)}")
print(f"  Unique URLs in xlsx      : {len(xlsx_urls)}")

# ── 2. S3 ─────────────────────────────────────────────────────────────────────
print("\nQuerying S3 ...")
session = boto3.Session(profile_name=os.getenv("AWS_PROFILE"))
s3 = session.client("s3", region_name=os.getenv("AWS_REGION"))
bucket = os.getenv("AWS_S3_BUCKET")
prefix = os.getenv("AWS_S3_PREFIX", "")

paginator = s3.get_paginator("list_objects_v2")
s3_keys: list[str] = []
s3_doc_ids: set[str] = set()
for page in paginator.paginate(Bucket=bucket, Prefix=prefix + "/"):
    for obj in page.get("Contents", []):
        key = obj["Key"]
        s3_keys.append(key)
        # Extract doc_id from key: <prefix>/<doc_id>/<file>
        parts = key.removeprefix(prefix).strip("/").split("/")
        if parts:
            s3_doc_ids.add(parts[0])

print(f"  S3 total objects         : {len(s3_keys)}")
print(f"  S3 unique document IDs   : {len(s3_doc_ids)}")

# ── 3. Qdrant ─────────────────────────────────────────────────────────────────
print("\nQuerying Qdrant ...")
qclient = QdrantClient(url=os.getenv("QDRANT_URL"), api_key=os.getenv("QDRANT_API_KEY"))
collection = os.getenv("QDRANT_COLLECTION")

qdrant_chunk_ids: set[str] = set()
qdrant_doc_ids: set[str] = set()
offset = None
while True:
    results, next_offset = qclient.scroll(
        collection_name=collection,
        limit=250,
        offset=offset,
        with_payload=True,
    )
    for r in results:
        cid = r.payload.get("storage_chunk_id") or r.payload.get("chunk_id")
        did = r.payload.get("document_id")
        if cid:
            qdrant_chunk_ids.add(str(cid))
        if did:
            qdrant_doc_ids.add(str(did))
    if next_offset is None:
        break
    offset = next_offset

print(f"  Qdrant total vectors     : {len(qdrant_chunk_ids)}")
print(f"  Qdrant unique doc IDs    : {len(qdrant_doc_ids)}")

# ── 4. Chunk ID comparison ────────────────────────────────────────────────────
print("\n=== Chunk ID Match (xlsx vs Qdrant) ===")
matched   = xlsx_chunk_ids & qdrant_chunk_ids
missing   = xlsx_chunk_ids - qdrant_chunk_ids   # in xlsx but NOT in Qdrant
extra     = qdrant_chunk_ids - xlsx_chunk_ids   # in Qdrant but NOT in xlsx

total = len(xlsx_chunk_ids)
pct   = 100 * len(matched) / total if total else 0
print(f"  xlsx chunk IDs           : {total}")
print(f"  Matched in Qdrant        : {len(matched)} ({pct:.1f}%)")
print(f"  Missing from Qdrant      : {len(missing)}")
print(f"  Extra in Qdrant (not in xlsx): {len(extra)}")

if missing:
    print("\nSample missing chunk IDs (first 10):")
    for cid in sorted(missing)[:10]:
        print(f"  {cid}")

# ── 5. Format analysis ────────────────────────────────────────────────────────
print("\n=== Chunk ID Format Analysis ===")
xlsx_fmt  = {c for c in xlsx_chunk_ids  if c.startswith("url_")}
q_fmt     = {c for c in qdrant_chunk_ids if c.startswith("url_")}
q_old_fmt = {c for c in qdrant_chunk_ids if not c.startswith("url_")}
print(f"  xlsx using url_* format  : {len(xlsx_fmt)} / {total}")
print(f"  Qdrant using url_* format: {len(q_fmt)} / {len(qdrant_chunk_ids)}")
if q_old_fmt:
    print(f"  Qdrant OLD format chunks : {len(q_old_fmt)} (should be 0 after re-upload)")
    print("  Sample old-format chunks:")
    for c in sorted(q_old_fmt)[:5]:
        print(f"    {c}")

print("\nDone.")
