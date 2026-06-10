"""Find which URLs the missing chunk IDs belong to in result.xlsx."""

import io
import sys

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")

from dotenv import load_dotenv  # noqa: E402

load_dotenv()

import openpyxl  # noqa: E402
from qdrant_client import QdrantClient  # noqa: E402

from agentic_rag.retrieval.config import require_qdrant_vector_store_config  # noqa: E402

XLSX_PATH = r"C:\Users\ACER\Downloads\Agentic_RAG_Group1\guide\reports\result.xlsx"
RELINK = r"C:\Users\ACER\Downloads\Agentic_RAG_Group1\_relink.txt"

# Load xlsx
wb = openpyxl.load_workbook(XLSX_PATH, read_only=True, data_only=True)
ws = wb.active
headers = [cell.value for cell in ws[2]]
col_id = headers.index("id")
col_chunk = headers.index("ground_truth_chunk_ids")
col_doc = headers.index("ground_truth_doc")

# Build: chunk_id -> (question_id, url)
chunk_to_row: dict[str, tuple] = {}
for row in ws.iter_rows(min_row=3, values_only=True):
    cid = row[col_chunk]
    url = row[col_doc]
    qid = row[col_id]
    if cid and str(cid).strip():
        chunk_to_row[str(cid).strip()] = (qid, str(url).strip() if url else "")

# Fetch Qdrant chunk IDs
qdrant_config = require_qdrant_vector_store_config()
assert qdrant_config.url is not None
qclient = QdrantClient(
    url=qdrant_config.url.get_secret_value(),
    api_key=(
        qdrant_config.api_key.get_secret_value() if qdrant_config.api_key is not None else None
    ),
)
collection = qdrant_config.collection
qdrant_chunk_ids: set[str] = set()
offset = None
while True:
    results, next_offset = qclient.scroll(
        collection_name=collection, limit=250, offset=offset, with_payload=True
    )
    for r in results:
        cid = r.payload.get("storage_chunk_id") or r.payload.get("chunk_id")
        if cid:
            qdrant_chunk_ids.add(str(cid))
    if next_offset is None:
        break
    offset = next_offset

xlsx_chunk_ids = set(chunk_to_row.keys())
missing = xlsx_chunk_ids - qdrant_chunk_ids

# Load relink.txt
relink_urls = set()
with open(RELINK, encoding="utf-8") as _f:
    for line in _f:
        line = line.strip()
        if line.startswith("http"):
            relink_urls.add(line)

print(f"Missing chunk IDs: {len(missing)}")
print()
url_status: dict[str, str] = {}
for cid in sorted(missing):
    qid, url = chunk_to_row[cid]
    in_relink = url in relink_urls
    url_status[url] = "IN_RELINK" if in_relink else "NOT_IN_RELINK"
    print(f"  chunk_id : {cid}")
    print(f"  question : {qid}")
    print(f"  url      : {url}")
    print(f"  relink   : {'YES' if in_relink else 'NO - was never uploaded'}")
    print()

print("=== URL summary ===")
url_missing = set(u for u, s in url_status.items())
print(f"Distinct URLs with missing chunks: {len(url_missing)}")
for url, st in sorted(url_status.items()):
    print(f"  [{st}] {url}")
